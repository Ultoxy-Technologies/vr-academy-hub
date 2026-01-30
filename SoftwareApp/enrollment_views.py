# views.py (enrollment section)
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.db.models import Q, Sum, Count, F
from django.core.paginator import Paginator
from django.utils import timezone
from datetime import datetime, date
from django.http import HttpResponse, JsonResponse
from django.contrib import messages
from decimal import Decimal
from AdminApp.models import *
from .forms import EnrollmentForm, BatchForm, EventForm, EventRegistrationForm, EventRegistrationBulkForm
# views.py
from django.db.models import Count, Sum, Avg, Q
from django.utils import timezone
from datetime import timedelta, datetime
from decimal import Decimal
import json

from functools import wraps

def has_a_auhtenticated_user(view_func):
    @wraps(view_func)
    def _wrapped_view(request, *args, **kwargs):
        user = request.user
        if user.is_authenticated and user.role == 'is_enrollment' or user.role == 'is_crm_and_enrollment':
            return view_func(request, *args, **kwargs)
        # if not student → redirect to home 
        if user.is_authenticated:
            if user.is_superuser:
                return redirect('/admin/')
            elif user.role=="is_crm_manager":
                return redirect('/software/crm_software_dashboard')
            elif user.role=="is_enrollment":
                return redirect('/software/enrollment-dashboard/')
            elif user.role=="is_crm_and_enrollment":
                return redirect('/software/software-welcome-page')
            elif user.role=="is_student":
                return redirect('/student/dashboard')
        else:
            messages.error(request,"Not sutable role found")
            return redirect('/')   # change 'home' if your url name is different
    return _wrapped_view



@login_required
@has_a_auhtenticated_user
def enrollment_dashboard(request):
    """
    Modern enrollment dashboard with comprehensive analytics
    """
    # Get filter parameters
    date_range = request.GET.get('date_range', 'month')  # day, week, month, year, custom
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    batch_filter = request.GET.get('batch', '')
    status_filter = request.GET.get('status', '')
    
    # Calculate date ranges
    today = timezone.now().date()
    
    if date_range == 'day':
        start_date_obj = today
        end_date_obj = today
    elif date_range == 'week':
        start_date_obj = today - timedelta(days=7)
        end_date_obj = today
    elif date_range == 'month':
        start_date_obj = today - timedelta(days=30)
        end_date_obj = today
    elif date_range == 'year':
        start_date_obj = today - timedelta(days=365)
        end_date_obj = today
    elif date_range == 'custom' and start_date and end_date:
        try:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
        except:
            start_date_obj = today - timedelta(days=30)
            end_date_obj = today
    else:
        start_date_obj = today - timedelta(days=30)
        end_date_obj = today
    
    # Base queryset
    enrollments = Enrollment.objects.filter(
        enrollment_date__range=[start_date_obj, end_date_obj]
    ).select_related('student', 'batch')
    
    # Apply additional filters
    if batch_filter:
        enrollments = enrollments.filter(batch_id=batch_filter)
    
    if status_filter:
        enrollments = enrollments.filter(status=status_filter)
    
    # Get all batches for filter dropdown
    batches = Batch.objects.filter(is_active=True).order_by('batch_code')
    
    # ========== KEY METRICS ==========
    total_enrollments = enrollments.count()
    
    # Enrollment by status
    active_enrollments = enrollments.filter(status='active').count()
    completed_enrollments = enrollments.filter(status='completed').count()
    pending_payments = enrollments.filter(payment_status__in=['pending', 'partial', 'overdue']).count()
    
    # Financial metrics - convert to float for template
    total_revenue = float(Decimal('0.00'))
    total_pending = float(Decimal('0.00'))
    total_collected = float(Decimal('0.00'))
    
    for enrollment in enrollments:
        total_revenue += float(enrollment.total_fees)
        total_collected += float(enrollment.paid_amount)
        total_pending += float(enrollment.balance)
    
    avg_fee_per_student = float(total_revenue / total_enrollments) if total_enrollments > 0 else 0.00
    
    # ========== CHARTS DATA ==========
    
    # 1. Enrollment Trend (Last 7 days)
    trend_data = []
    for i in range(7):
        date = today - timedelta(days=6-i)
        count = Enrollment.objects.filter(
            enrollment_date=date
        ).count()
        trend_data.append({
            'date': date.strftime('%d %b'),
            'count': count
        })
    
    # 2. Batch-wise Enrollment Distribution
    batch_distribution = []
    batch_wise_data = enrollments.values('batch__batch_code', 'batch__batch_title').annotate(
        count=Count('id'),
        total_fees=Sum('total_fees'),
        collected=Sum('payments__amount')
    ).order_by('-count')
    
    for item in batch_wise_data[:10]:  # Top 10 batches
        batch_distribution.append({
            'batch': f"{item['batch__batch_code'] or 'N/A'} - {(item['batch__batch_title'] or '')[:20]}",
            'enrollments': item['count'],
            'revenue': float(item['total_fees'] or 0),
            'collected': float(item['collected'] or 0)
        })
    
    # 3. Status Distribution
    status_distribution = []
    status_data = enrollments.values('status').annotate(count=Count('id'))
    for item in status_data:
        status_distribution.append({
            'status': item['status'].title(),
            'count': item['count']
        })
    
    # 4. Payment Status Distribution
    payment_status_data = []
    payment_stats = enrollments.values('payment_status').annotate(count=Count('id'))
    for item in payment_stats:
        payment_status_data.append({
            'status': item['payment_status'].title(),
            'count': item['count']
        })
    
    # 5. Monthly Revenue Trend
    monthly_revenue = []
    current_year = today.year
    for month in range(1, 13):
        month_start = datetime(current_year, month, 1).date()
        if month == 12:
            month_end = datetime(current_year + 1, 1, 1).date() - timedelta(days=1)
        else:
            month_end = datetime(current_year, month + 1, 1).date() - timedelta(days=1)
        
        month_enrollments = Enrollment.objects.filter(
            enrollment_date__range=[month_start, month_end]
        )
        
        month_revenue = Decimal('0.00')
        month_collected = Decimal('0.00')
        
        for enrollment in month_enrollments:
            month_revenue += enrollment.total_fees
            month_collected += enrollment.paid_amount
        
        monthly_revenue.append({
            'month': month_start.strftime('%b'),
            'revenue': float(month_revenue),
            'collected': float(month_collected)
        })
    
    # ========== RECENT ACTIVITIES ==========
    recent_enrollments = Enrollment.objects.select_related(
        'student', 'batch'
    ).order_by('-enrollment_date')[:10]
    
    recent_payments = Payment.objects.select_related(
        'enrollment', 'enrollment__student'
    ).order_by('-payment_date')[:10]
    
    # ========== BATCH PERFORMANCE ==========
     
    
    # Helper function to convert Decimal to float for JSON serialization
    def decimal_to_float(obj):
        if isinstance(obj, Decimal):
            return float(obj)
        elif isinstance(obj, dict):
            return {k: decimal_to_float(v) for k, v in obj.items()}
        elif isinstance(obj, list):
            return [decimal_to_float(item) for item in obj]
        return obj
    
    # Prepare context
    context = {
        # Filters
        'batches': batches,
        'status_choices': Enrollment.STATUS_CHOICES,
        'date_range': date_range,
        'start_date': start_date,
        'end_date': end_date,
        'batch_filter': batch_filter,
        'status_filter': status_filter,
        
        # Key Metrics (already converted to float)
        'total_enrollments': total_enrollments,
        'active_enrollments': active_enrollments,
        'completed_enrollments': completed_enrollments,
        'pending_payments': pending_payments,
        'total_revenue': total_revenue,
        'total_collected': total_collected,
        'total_pending': total_pending,
        'avg_fee_per_student': avg_fee_per_student,
        
        # Chart Data (JSON for JavaScript) - convert all to float
        'trend_data_json': json.dumps(trend_data),
        'batch_distribution_json': json.dumps(batch_distribution),
        'status_distribution_json': json.dumps(status_distribution),
        'payment_status_json': json.dumps(payment_status_data),
        'monthly_revenue_json': json.dumps(monthly_revenue),
        
        # Recent Activities
        'recent_enrollments': recent_enrollments,
        'recent_payments': recent_payments,
         
        
        # Date info
        'start_date_obj': start_date_obj.strftime('%Y-%m-%d'),
        'end_date_obj': end_date_obj.strftime('%Y-%m-%d'),
        'today': today.strftime('%Y-%m-%d'),
    }
    
    return render(request, 'enrollment_dashboard.html', context)



# views.py
import pandas as pd
from django.http import HttpResponse
from io import BytesIO
from datetime import datetime, timedelta
import json


@login_required
@has_a_auhtenticated_user
def export_enrollments_filtered(request):
    """
    Export filtered enrollments to Excel with proper formatting
    """
    # Get all filter parameters from request
    search_query = request.GET.get('search', '')
    status_filter = request.GET.get('status', '')
    batch_filter = request.GET.get('batch', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    payment_status_filter = request.GET.get('payment_status', '')
    
    # Apply filters (same logic as enrolled_student_list)
    enrollments = Enrollment.objects.all().select_related(
        'student', 'batch', 'batch__basic_to_advance_cource', 'batch__advance_to_pro_cource'
    ).prefetch_related('payments').order_by('-enrollment_date')
    
    if search_query:
        enrollments = enrollments.filter(
            Q(student__name__icontains=search_query) |
            Q(student__mobile_number__icontains=search_query) |
            Q(course_title__icontains=search_query)
        )
    
    if status_filter:
        enrollments = enrollments.filter(status=status_filter)
    
    if batch_filter:
        enrollments = enrollments.filter(batch_id=batch_filter)
    
    if payment_status_filter:
        enrollments = enrollments.filter(payment_status=payment_status_filter)
    
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d')
            enrollments = enrollments.filter(enrollment_date__gte=date_from_obj)
        except ValueError:
            pass
    
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d')
            enrollments = enrollments.filter(enrollment_date__lte=date_to_obj)
        except ValueError:
            pass
    
    # Prepare data for Excel
    data = []
    
    for enrollment in enrollments:
        # Get batch details
        batch_info = ""
        if enrollment.batch:
            batch_info = f"{enrollment.batch.batch_code} - {enrollment.batch.batch_title}"
            if enrollment.batch.course:
                batch_info += f" ({enrollment.batch.course.title})"
        
        # Get payment details
        payment_details = []
        for payment in enrollment.payments.all():
            payment_details.append(
                f"{payment.payment_date.strftime('%d-%m-%Y')}: ₹{payment.amount} "
                f"({payment.get_payment_mode_display()})"
            )
        
        # Get certificate info
        certificate_info = ""
        if hasattr(enrollment, 'certificate'):
            certificate_info = f"{enrollment.certificate.certificate_number} - {enrollment.certificate.issue_date}"
        
        # Calculate payment summary
        paid_amount = float(enrollment.paid_amount)
        balance_amount = float(enrollment.balance)
        payment_progress = float(enrollment.payment_progress)
        
        # Prepare row data
        row = {
            'Enrollment ID': f"ENR-{str(enrollment.pk).zfill(6)}",
            'Enrollment Date': enrollment.enrollment_date.strftime('%d-%m-%Y'),
            
            # Student Information
            'Student ID': enrollment.student.id if enrollment.student else '',
            'Student Name': enrollment.student.name if enrollment.student else '',
            'Mobile Number': enrollment.student.mobile_number if enrollment.student else '',
            'Email': enrollment.student.email if enrollment.student else '',
            'Address': enrollment.student.address if hasattr(enrollment.student, 'address') else '',
            
            # Course & Batch Information
            'Batch Code': enrollment.batch.batch_code if enrollment.batch else '',
            'Batch Title': enrollment.batch.batch_title if enrollment.batch else '',
            'Course Type': enrollment.batch.course_type if enrollment.batch else '',
            'Course Title': enrollment.course_title,
            'Batch Start Date': enrollment.batch.start_date.strftime('%d-%m-%Y') if enrollment.batch and enrollment.batch.start_date else '',
            'Batch End Date': enrollment.batch.end_date.strftime('%d-%m-%Y') if enrollment.batch and enrollment.batch.end_date else '',
            'Class Start Date': enrollment.class_start_date.strftime('%d-%m-%Y') if enrollment.class_start_date else '',
            'Class End Date': enrollment.class_end_date.strftime('%d-%m-%Y') if enrollment.class_end_date else '',
            
            # Enrollment Details
            'Enrollment Status': enrollment.get_status_display(),
            'Payment Status': enrollment.get_payment_status_display(),
            'Payment Schedule': enrollment.payment_schedule,
            
            # Financial Information
            'Total Fees': float(enrollment.total_fees),
            'Discount': float(enrollment.discount),
            'Net Fees': float(enrollment.net_fees),
            'Paid Amount': paid_amount,
            'Balance Amount': balance_amount,
            'Payment Progress %': payment_progress,
            
            # Payment Details (comma separated)
            'Payment History': ' | '.join(payment_details) if payment_details else 'No Payments',
            'Total Payments': enrollment.payments.count(),
            'Last Payment Date': enrollment.last_payment.payment_date.strftime('%d-%m-%Y') if enrollment.last_payment else '',
            'Last Payment Amount': float(enrollment.last_payment.amount) if enrollment.last_payment else 0,
            
            # Certificate Information
            'Certificate Number': certificate_info,
            'Certificate Issued': 'Yes' if hasattr(enrollment, 'certificate') else 'No',
            'Certificate Issue Date': enrollment.certificate.issue_date.strftime('%d-%m-%Y') if hasattr(enrollment, 'certificate') else '',
            
            # Additional Information
            'Days Since Enrollment': (datetime.now().date() - enrollment.enrollment_date).days,
            'Is Payment Overdue': 'Yes' if (balance_amount > 0 and enrollment.class_start_date and enrollment.class_start_date <= datetime.now().date()) else 'No',
            'Created At': enrollment.created_at.strftime('%d-%m-%Y %H:%M:%S') if hasattr(enrollment, 'created_at') else '',
            'Updated At': enrollment.updated_at.strftime('%d-%m-%Y %H:%M:%S') if hasattr(enrollment, 'updated_at') else '',
        }
        
        data.append(row)
    
    # Create DataFrame
    df = pd.DataFrame(data)
    
    # Create Excel file in memory
    output = BytesIO()
    
    # Create Excel writer
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Write main data
        df.to_excel(writer, sheet_name='Enrollments', index=False)
        
        # Auto-adjust columns width
        worksheet = writer.sheets['Enrollments']
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Create summary sheet
        summary_data = {
            'Metric': [
                'Total Enrollments',
                'Active Enrollments',
                'Completed Enrollments',
                'Total Revenue',
                'Total Collected',
                'Total Pending',
                'Average Fee per Student',
                'Pending Payments Count'
            ],
            'Value': [
                len(df),
                len(df[df['Enrollment Status'] == 'Active']),
                len(df[df['Enrollment Status'] == 'Completed']),
                df['Total Fees'].sum(),
                df['Paid Amount'].sum(),
                df['Balance Amount'].sum(),
                df['Total Fees'].mean() if len(df) > 0 else 0,
                len(df[df['Payment Status'].isin(['Pending', 'Partial', 'Overdue'])])
            ]
        }
        
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
        
        # Auto-adjust summary columns
        summary_worksheet = writer.sheets['Summary']
        for column in summary_worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            summary_worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Create batch-wise summary
        if 'Batch Code' in df.columns and not df['Batch Code'].isnull().all():
            batch_summary = df.groupby('Batch Code').agg({
                'Student Name': 'count',
                'Total Fees': 'sum',
                'Paid Amount': 'sum',
                'Balance Amount': 'sum'
            }).reset_index()
            
            batch_summary.columns = ['Batch Code', 'Total Students', 'Total Fees', 'Amount Collected', 'Amount Pending']
            batch_summary.to_excel(writer, sheet_name='Batch Summary', index=False)
            
            # Auto-adjust batch summary columns
            batch_worksheet = writer.sheets['Batch Summary']
            for column in batch_worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                batch_worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Prepare HTTP response
    output.seek(0)
    
    # Create filename with timestamp and filters
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filters = []
    
    if search_query:
        filters.append(f"search_{search_query}")
    if status_filter:
        filters.append(f"status_{status_filter}")
    if batch_filter:
        filters.append(f"batch_{batch_filter}")
    if date_from:
        filters.append(f"from_{date_from}")
    if date_to:
        filters.append(f"to_{date_to}")
    
    filter_str = '_'.join(filters) if filters else 'all'
    filename = f"enrollments_export_{filter_str}_{timestamp}.xlsx"
    
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response




@login_required
@has_a_auhtenticated_user
def enrolled_student_list(request):
    """
    List all enrollments with filtering, pagination, and statistics
    Follows the exact pattern of crm_follow_up_list
    """
    # Get filter parameters
    search_query = request.GET.get('search', '')
    status_filter = request.GET.get('status', '')
    batch_filter = request.GET.get('batch', '')
    branch_filter = request.GET.get('branch', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    payment_status_filter = request.GET.get('payment_status', '')
    
    # Start with all enrollments
    enrollments = Enrollment.objects.all().select_related(
        'student', 'batch', 'batch__basic_to_advance_cource', 'batch__advance_to_pro_cource'
    ).prefetch_related('payments').order_by('-enrollment_date')
    
    # Get all batches for filter dropdown
    batches = Batch.objects.filter(is_active=True).order_by('batch_code')
    branches = Branch.objects.filter().order_by('branch_name')
    
    # Apply filters
    if search_query:
        enrollments = enrollments.filter(
            Q(student__name__icontains=search_query) |
            Q(student__mobile_number__icontains=search_query) |
            Q(course_title__icontains=search_query)
        )
    
    if status_filter:
        enrollments = enrollments.filter(status=status_filter)
    
    if batch_filter:
        enrollments = enrollments.filter(batch_id=batch_filter)
    
    if branch_filter:
        enrollments = enrollments.filter(branch__id=branch_filter)

    if payment_status_filter:
        enrollments = enrollments.filter(payment_status=payment_status_filter)
    
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d')
            enrollments = enrollments.filter(enrollment_date__gte=date_from_obj)
        except ValueError:
            pass
    
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d')
            enrollments = enrollments.filter(enrollment_date__lte=date_to_obj)
        except ValueError:
            pass
    
    # Calculate statistics on ALL filtered data (not just paginated)
    total_enrollments = enrollments.count()
    active_enrollments = enrollments.filter(status='active').count()
    completed_enrollments = enrollments.filter(status='completed').count()
    pending_payments = enrollments.filter(payment_status__in=['pending', 'partial', 'overdue']).count()
    
    # Today's enrollments - FIXED
    today = date.today()
    today_enrollments = enrollments.filter(enrollment_date=today).count()
    
    # Pagination - DO THIS BEFORE calculating revenue for each enrollment
    paginator = Paginator(enrollments, 50)  # Show 10 enrollments per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Calculate payment data for each enrollment in the CURRENT PAGE only
    total_revenue = Decimal('0.00')
    pending_revenue = Decimal('0.00')
    
    for enrollment in page_obj:  # Use page_obj, not enrollments
        # Calculate paid amount from prefetched payments
        paid_amount = sum(payment.amount for payment in enrollment.payments.all())
        enrollment.display_paid = paid_amount
        enrollment.display_balance = max(enrollment.net_fees - paid_amount, Decimal('0.00'))
        
        # Calculate progress
        if enrollment.net_fees > 0:
            enrollment.display_progress = min((paid_amount / enrollment.net_fees * 100), 100)
        else:
            enrollment.display_progress = 100
        
        # Days since enrollment
        enrollment.days_since_enrollment = (today - enrollment.enrollment_date).days
        
        # Check if enrollment is overdue
        enrollment.is_payment_overdue = (
            enrollment.display_balance > 0 and 
            enrollment.class_start_date and 
            enrollment.class_start_date <= today
        )
        
        # Generate enrollment ID for display
        enrollment.enrollment_id_display = f"ENR-{str(enrollment.pk).zfill(6)}"
        
        # Add to totals (for current page only)
        total_revenue += paid_amount
        pending_revenue += enrollment.display_balance
    
    # Calculate overall revenue totals (for all filtered data)
    overall_total_revenue = Decimal('0.00')
    overall_pending_revenue = Decimal('0.00')
    
    # If you want to calculate overall totals, you need to do a separate query
    # Or add these calculations in a different way
    
    # Prepare context
    context = {
        'enrollments': page_obj,  # Use page_obj instead of enrollments
        'batches': batches,
        'branches': branches,
        
        # Statistics
        'total_enrollments': total_enrollments,
        'active_enrollments': active_enrollments,
        'completed_enrollments': completed_enrollments,
        'pending_payments': pending_payments,
        'today_enrollments': today_enrollments,
        'total_revenue': total_revenue,  # This is for current page only
        'pending_revenue': pending_revenue,  # This is for current page only
        
        # Filter values
        'search_query': search_query,
        'status_filter': status_filter,
        'batch_filter': batch_filter,
        'payment_status_filter': payment_status_filter,
        'branch_filter': branch_filter,
        'date_from': date_from,
        'date_to': date_to,
        
        # Status choices
        'status_choices': Enrollment.STATUS_CHOICES,
        'payment_status_choices': Enrollment.PAYMENT_STATUS_CHOICES,
        
        # Pagination context
        'page_obj': page_obj,
    }
    
    return render(request, 'enrolled_student_list.html', context)


# views.py
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.db.models import Sum
from django.utils import timezone
from decimal import Decimal 
from django import forms


@login_required
@has_a_auhtenticated_user
def enrollment_detail(request, pk):
    """View enrollment details with payment history and certificate info"""
    enrollment = get_object_or_404(
        Enrollment.objects.select_related(
            'student', 'batch', 'batch__basic_to_advance_cource', 'batch__advance_to_pro_cource'
        ).prefetch_related('payments', 'payments__received_by'),
        pk=pk
    )
    
    # Get all payments for this enrollment
    payments = enrollment.payments.all().order_by('-payment_date')
    
    # Calculate payment statistics
    total_paid = payments.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    total_due = enrollment.balance
    
    # Get payment summary by type
    payment_by_type = {}
    for payment_type, display_name in Payment.PAYMENT_TYPE_CHOICES:
        amount = payments.filter(payment_type=payment_type).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
        if amount > 0:
            payment_by_type[display_name] = amount
    
    # Get certificate if exists
    certificate = None
    try:
        certificate = StudentCertificate.objects.get(enrollment=enrollment)
    except StudentCertificate.DoesNotExist:
        certificate = None
    
    # Calculate days since enrollment
    days_since_enrollment = (timezone.now().date() - enrollment.enrollment_date).days
    
    # Calculate class progress percentage
    class_progress = 0
    if enrollment.class_start_date and enrollment.class_end_date:
        total_days = (enrollment.class_end_date - enrollment.class_start_date).days
        if total_days > 0:
            days_passed = (timezone.now().date() - enrollment.class_start_date).days
            class_progress = min(max((days_passed / total_days) * 100, 0), 100)
    
    # Check if enrollment is overdue
    is_overdue = (
        total_due > 0 and 
        enrollment.class_start_date and 
        enrollment.class_start_date <= timezone.now().date()
    )
    
    # Get batch details
    batch = enrollment.batch
    batch_details = {
        'start_date': batch.start_date,
        'end_date': batch.end_date,
        'is_active': batch.is_active,
        'course_type': batch.course_type,
        'course_fees': batch.course_fees,
    }
    
    # Payment form for quick payment recording
    # payment_form = PaymentForm()
    # payment_form.fields['enrollment'].initial = enrollment
    # payment_form.fields['enrollment'].widget = forms.HiddenInput()
    
    context = {
        'enrollment': enrollment,
        'payments': payments,
        'total_paid': total_paid,
        'total_due': total_due,
        'payment_by_type': payment_by_type,
        'certificate': certificate,
        'days_since_enrollment': days_since_enrollment,
        'class_progress': class_progress,
        'is_overdue': is_overdue,
        'batch_details': batch_details,
        # 'payment_form': payment_form,
        
        # For progress bars
        'payment_progress': enrollment.payment_progress,
        'total_installments': enrollment.total_installments if hasattr(enrollment, 'total_installments') else 0,
        'paid_installments': enrollment.paid_installments if hasattr(enrollment, 'paid_installments') else 0,
    }
    
    return render(request, 'enrollment_details.html', context)





@login_required
@has_a_auhtenticated_user
def export_enrollments(request):
    """
    Export enrollments to CSV (following export_followups pattern)
    """
    import csv
    
    # Create the HttpResponse object with CSV header
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="enrollments_export_{}.csv"'.format(
        date.today().strftime('%Y%m%d')
    )
    
    writer = csv.writer(response)
    
    # Write header row
    writer.writerow([
        'Enrollment ID', 'Student Name', 'Mobile Number', 'Batch',
        'Course', 'Enrollment Date', 'Status', 'Payment Status',
        'Total Fees', 'Discount', 'Net Fees', 'Paid Amount', 'Balance',
        'Class Start Date', 'Class End Date', 'Created At'
    ])
    
    # Get all enrollments with related data
    enrollments = Enrollment.objects.all().select_related(
        'student', 'batch'
    )
    
    # Write data rows
    for enrollment in enrollments:
        # Get course title from batch
        course_title = "Unknown"
        if enrollment.batch:
            if enrollment.batch.basic_to_advance_cource:
                course_title = enrollment.batch.basic_to_advance_cource.title
            elif enrollment.batch.advance_to_pro_cource:
                course_title = enrollment.batch.advance_to_pro_cource.title
        
        writer.writerow([
            enrollment.enrollment_id,
            enrollment.student.name if enrollment.student else '',
            enrollment.student.mobile_number if enrollment.student else '',
            enrollment.batch.batch_title if enrollment.batch else '',
            course_title,
            enrollment.enrollment_date.strftime('%Y-%m-%d'),
            enrollment.get_status_display(),
            enrollment.get_payment_status_display(),
            str(enrollment.total_fees),
            str(enrollment.discount),
            str(enrollment.net_fees),
            str(enrollment.paid_amount),
            str(enrollment.balance),
            enrollment.class_start_date.strftime('%Y-%m-%d') if enrollment.class_start_date else '',
            enrollment.class_end_date.strftime('%Y-%m-%d') if enrollment.class_end_date else '',
            enrollment.created_at.strftime('%Y-%m-%d %H:%M') if hasattr(enrollment, 'created_at') else '',
        ])
    
    return response



@login_required
@has_a_auhtenticated_user
def download_enrollment_template(request):
    """
    Download template CSV file for enrollment import
    """
    import csv
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="enrollment_template.csv"'
    
    writer = csv.writer(response)
    
    # Write header row with example data
    writer.writerow([
        'Student Name', 'Mobile Number', 'Batch Code', 
        'Total Fees', 'Discount', 'Status', 'Payment Status'
    ])
    
    # Example rows
    writer.writerow([
        'John Doe', '9876543210', 'BATCH-2024-001',
        '15000.00', '1000.00', 'active', 'pending'
    ])
    
    writer.writerow([
        'Jane Smith', '9876543211', 'BATCH-2024-002',
        '20000.00', '0.00', 'active', 'partial'
    ])
    
    return response

import csv
from decimal import Decimal
from .forms import PaymentForm


@login_required
@has_a_auhtenticated_user
def import_enrollments(request):
    """
    Import enrollments from CSV (following import_followups pattern)
    """
    if request.method == 'POST' and request.FILES.get('csv_file'):
        csv_file = request.FILES['csv_file']
        
        # Check if file is CSV
        if not csv_file.name.endswith('.csv'):
            messages.error(request, 'Please upload a CSV file')
            return redirect('enrolled_student_list')
        
        try:
            # Read CSV file
            decoded_file = csv_file.read().decode('utf-8').splitlines()
            reader = csv.DictReader(decoded_file)
            
            imported_count = 0
            errors = []
            
            for row_num, row in enumerate(reader, start=2):  # start=2 for header row
                try:
                    # Get student by mobile number or name
                    mobile_number = row.get('Mobile Number', '').strip()
                    student_name = row.get('Student Name', '').strip()
                    
                    # Find or create student
                    student = None
                    if mobile_number:
                        student = CustomUser.objects.filter(mobile_number=mobile_number).first()
                    
                    if not student and student_name:
                        # Try to find by name
                        student = CustomUser.objects.filter(name__icontains=student_name).first()
                    
                    if not student:
                        errors.append(f"Row {row_num}: Student not found ({student_name} - {mobile_number})")
                        continue
                    
                    # Get batch
                    batch_code = row.get('Batch Code', '').strip()
                    batch = None
                    if batch_code:
                        batch = Batch.objects.filter(batch_code=batch_code).first()
                    
                    if not batch:
                        errors.append(f"Row {row_num}: Batch not found ({batch_code})")
                        continue
                    
                    # Parse numeric values
                    try:
                        total_fees = Decimal(row.get('Total Fees', '0').strip())
                        discount = Decimal(row.get('Discount', '0').strip())
                    except:
                        errors.append(f"Row {row_num}: Invalid fee/discount format")
                        continue
                    
                    # Create enrollment
                    Enrollment.objects.create(
                        student=student,
                        batch=batch,
                        total_fees=total_fees,
                        discount=discount,
                        status=row.get('Status', 'active').strip(),
                        payment_status=row.get('Payment Status', 'pending').strip(),
                        enrollment_date=timezone.now().date(),
                    )
                    
                    imported_count += 1
                    
                except Exception as e:
                    errors.append(f"Row {row_num}: {str(e)}")
            
            # Show results
            if imported_count > 0:
                messages.success(request, f'Successfully imported {imported_count} enrollments')
            
            if errors:
                error_list = "<br>".join(errors[:10])  # Show first 10 errors
                if len(errors) > 10:
                    error_list += f"<br>... and {len(errors) - 10} more errors"
                messages.warning(request, f'Import completed with errors:<br>{error_list}')
        
        except Exception as e:
            messages.error(request, f'Error importing file: {str(e)}')
    
    return redirect('enrolled_student_list')



@login_required
@has_a_auhtenticated_user
def record_payment(request, enrollment_id):
    """Record a new payment for enrollment"""
    enrollment = get_object_or_404(Enrollment, id=enrollment_id)
    
    # Calculate payment statistics with safe defaults
    payment_stats = {
        'total_fees': enrollment.total_fees or Decimal('0.00'),
        'discount': enrollment.discount or Decimal('0.00'),
        'net_fees': enrollment.net_fees or Decimal('0.00'),
        'paid_amount': enrollment.paid_amount,
        'balance': enrollment.balance,
        'payment_progress': enrollment.payment_progress or 0,
    }
    
    if request.method == 'POST':
        form = PaymentForm(request.POST, enrollment=enrollment, request=request)
        
        if form.is_valid():
            try:
                payment = form.save()
                
                # Update payment status message based on balance
                new_balance = enrollment.balance
                if new_balance <= 0:
                    status_msg = "Fully Paid! All fees have been cleared."
                else:
                    status_msg = f"Payment recorded. Remaining balance: ₹{new_balance}"
                
                messages.success(
                    request, 
                    f'Payment of ₹{payment.amount} recorded successfully for {enrollment.student.name}. {status_msg}'
                )
                
                # Return JavaScript to close window
                return HttpResponse("""
                    <script>
                        if (window.opener && !window.opener.closed) {
                            window.opener.location.reload();
                        }
                        window.close();
                    </script>
                """)
            except Exception as e:
                messages.error(request, f'Error recording payment: {str(e)}')
        else:
            # Collect all form errors
            error_messages = []
            for field, errors in form.errors.items():
                field_name = field.replace('_', ' ').title()
                for error in errors:
                    error_messages.append(f"{field_name}: {error}")
            
            if error_messages:
                messages.error(request, '<br>'.join(error_messages))
    
    else:
        form = PaymentForm(enrollment=enrollment, request=request)
    
    # Get payment history
    payment_history = enrollment.payments.all().order_by('-payment_date')
    
    context = {
        'form': form,
        'enrollment': enrollment,
        'payment_stats': payment_stats,
        'payment_history': payment_history,
    }
    
    return render(request, 'record_payment.html', context)


@login_required
@has_a_auhtenticated_user
def update_payment(request, id):
    """Update existing payment"""
    payment = get_object_or_404(Payment, id=id)
    enrollment = payment.enrollment
    
    if request.method == 'POST':
        form = PaymentForm(request.POST, instance=payment, enrollment=enrollment, request=request)
        
        if form.is_valid():
            try:
                form.save()
                messages.success(request, f'Payment updated successfully')
                
                # Return JavaScript to close window
                return HttpResponse("""
                    <script>
                        if (window.opener && !window.opener.closed) {
                            window.opener.location.reload();
                        }
                        window.close();
                    </script>
                """)
            except Exception as e:
                messages.error(request, f'Error updating payment: {str(e)}')
        else:
            # Collect all form errors
            error_messages = []
            for field, errors in form.errors.items():
                field_name = field.replace('_', ' ').title()
                for error in errors:
                    error_messages.append(f"{field_name}: {error}")
            
            if error_messages:
                messages.error(request, '<br>'.join(error_messages))
    
    else:
        form = PaymentForm(instance=payment, enrollment=enrollment, request=request)
    
    context = {
        'form': form,
        'payment': payment,
        'enrollment': enrollment,
    }
    
    return render(request, 'update_payment.html', context)


@login_required
@has_a_auhtenticated_user
def delete_payment(request, id):
    """Delete payment with confirmation"""
    payment = get_object_or_404(Payment, id=id)
    
    amount = payment.amount
    student_name = payment.enrollment.student.name
    payment.delete()
    
    # Update enrollment payment status
    payment.enrollment.update_payment_status()
    payment.enrollment.save()
    
    messages.success(request, f'Payment of ₹{amount} deleted for {student_name}')
    return redirect(request.META.get('HTTP_REFERER', 'payment_list'))    


# views.py
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.contrib import messages
from .forms import EnrollmentForm


@login_required
@has_a_auhtenticated_user
def create_enrollment(request):
    """Create new enrollment (opens in modal window)"""
    user=CustomUser.objects.filter(role="is_student")
    student_list=user.values_list('name','mobile_number')
    print(student_list)

    if request.method == 'POST':
        form = EnrollmentForm(request.POST, request=request)
        
        if form.is_valid():
            try:
                enrollment = form.save(commit=True)
                messages.success(request, f'Enrollment created successfully for {enrollment.student.name}')
                
                # Return JavaScript to close window
                return HttpResponse("""
                    <script>
                        if (window.opener && !window.opener.closed) {
                            window.opener.location.reload();
                        }
                        window.close();
                    </script>
                """)
            except Exception as e:
                messages.error(request, f'Error creating enrollment: {str(e)}')
        else:
            # Collect all form errors
            error_messages = []
            for field, errors in form.errors.items():
                field_name = field.replace('_', ' ').title()
                for error in errors:
                    error_messages.append(f"{field_name}: {error}")
            
            if error_messages:
                messages.error(request, '<br>'.join(error_messages))
    
    else:
        form = EnrollmentForm(request=request)
    
    return render(request, 'enrollment_create.html', {'form': form})


@login_required
@has_a_auhtenticated_user
def update_enrollment(request, pk):
    """Update existing enrollment (opens in modal window)"""
    enrollment = get_object_or_404(Enrollment, pk=pk)
    
    if request.method == 'POST':
        form = EnrollmentForm(request.POST, instance=enrollment, request=request, is_update=True)
        
        if form.is_valid():
            try:
                enrollment = form.save(commit=True)
                messages.success(request, f'Enrollment updated successfully for {enrollment.student.name}')
                
                # Return JavaScript to close window
                return HttpResponse("""
                    <script>
                        if (window.opener && !window.opener.closed) {
                            window.opener.location.reload();
                        }
                        window.close();
                    </script>
                """)
            except Exception as e:
                messages.error(request, f'Error updating enrollment: {str(e)}')
        else:
            # Collect all form errors
            error_messages = []
            for field, errors in form.errors.items():
                field_name = field.replace('_', ' ').title()
                for error in errors:
                    error_messages.append(f"{field_name}: {error}")
            
            if error_messages:
                messages.error(request, '<br>'.join(error_messages))
    
    else:
        form = EnrollmentForm(instance=enrollment, request=request, is_update=True)
    
    return render(request, 'enrollment_update.html', {'form': form, 'enrollment': enrollment})


@login_required
@has_a_auhtenticated_user
def delete_enrollment(request, pk):
    """Delete enrollment with confirmation"""
    enrollment = get_object_or_404(Enrollment, pk=pk)
    print(enrollment)
    
    if enrollment:
        student_name = enrollment.student.name
        enrollment.delete()
        messages.success(request, f'Enrollment for {student_name} deleted successfully')
        return redirect('/software/enrollments/')
    
    # For GET request, should be handled by JavaScript modal
    return redirect('/software/enrollments/')












from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from datetime import datetime, date
from decimal import Decimal
from django.db.models import Q, Sum


@login_required
@has_a_auhtenticated_user
def batch_list(request):
    """
    List all batches with filtering, pagination and statistics
    """
    # Get filter parameters
    search_query = request.GET.get('search', '')
    status_filter = request.GET.get('status', '')
    course_type_filter = request.GET.get('course_type', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    active_filter = request.GET.get('is_active', '')
    
    # Start with all batches
    batches = Batch.objects.all().order_by('-start_date')
    
    # Apply filters
    if search_query:
        batches = batches.filter(
            Q(batch_code__icontains=search_query) |
            Q(batch_title__icontains=search_query) |
            Q(description__icontains=search_query)
        )
    
    if status_filter:
        batches = batches.filter(batch_status=status_filter)
    
    if course_type_filter:
        if course_type_filter == 'basic':
            batches = batches.filter(basic_to_advance_cource__isnull=False)
        elif course_type_filter == 'advance':
            batches = batches.filter(advance_to_pro_cource__isnull=False)
    
    if active_filter:
        if active_filter == 'active':
            batches = batches.filter(is_active=True)
        elif active_filter == 'inactive':
            batches = batches.filter(is_active=False)
    
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d')
            batches = batches.filter(start_date__gte=date_from_obj)
        except ValueError:
            pass
    
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d')
            batches = batches.filter(start_date__lte=date_to_obj)
        except ValueError:
            pass
    
    # Calculate statistics on ALL filtered data (before pagination)
    total_batches = batches.count()
    active_batches = batches.filter(is_active=True).count()
    upcoming_batches = batches.filter(batch_status='upcoming').count()
    ongoing_batches = batches.filter(batch_status='ongoing').count()
    completed_batches = batches.filter(batch_status='completed').count()
    
    # Pagination (10 items per page)
    paginator = Paginator(batches, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Add calculated properties to each batch in CURRENT PAGE
    total_students_current_page = 0
    total_revenue_current_page = Decimal('0.00')
    
    for batch in page_obj:
        batch.display_current_students = batch.current_students
        batch.display_available_seats = batch.available_seats
        batch.display_occupancy = batch.occupancy_percentage
        batch.display_can_enroll = batch.can_enroll
        batch.display_duration = batch.duration_days
        batch.display_days_remaining = batch.days_remaining
        
        # Get course fees
        batch.display_fees = batch.course_fees
        
        # Get course title
        if batch.basic_to_advance_cource:
            batch.display_course_title = batch.basic_to_advance_cource.title
        elif batch.advance_to_pro_cource:
            batch.display_course_title = batch.advance_to_pro_cource.title
        else:
            batch.display_course_title = "No Course Assigned"
        
        # Add to page totals
        total_students_current_page += batch.current_students
        
        # Calculate batch revenue (if you have this method)
        try:
            batch_revenue = batch.total_revenue()
            batch.display_revenue = batch_revenue
            total_revenue_current_page += batch_revenue
        except:
            batch.display_revenue = Decimal('0.00')
    
    # For overall totals (all filtered batches), you might want to calculate efficiently
    # Using aggregate queries
    overall_total_students = sum(batch.current_students for batch in batches)
    
    # Prepare context
    context = {
        'batches': page_obj,
        
        # Statistics
        'total_batches': total_batches,
        'active_batches': active_batches,
        'upcoming_batches': upcoming_batches,
        'ongoing_batches': ongoing_batches,
        'completed_batches': completed_batches,
        'total_students': overall_total_students,  # All filtered batches
        'total_revenue': total_revenue_current_page,  # Current page only
        
        # Filter values
        'search_query': search_query,
        'status_filter': status_filter,
        'course_type_filter': course_type_filter,
        'date_from': date_from,
        'date_to': date_to,
        'active_filter': active_filter,
        
        # Choices for filters
        'status_choices': Batch.BATCH_STATUS_CHOICES,
        'course_type_choices': [
            ('', 'All Types'),
            ('basic', 'Basic to Advance'),
            ('advance', 'Advance to Pro'),
        ],
        'active_choices': [
            ('', 'All'),
            ('active', 'Active Only'),
            ('inactive', 'Inactive Only'),
        ],
        
        # Pagination context
        'page_obj': page_obj,
    }
    
    return render(request, 'batch_list.html', context)



@login_required
@has_a_auhtenticated_user
def create_batch(request):
    """Create new batch (opens in modal window)"""
    if request.method == 'POST':
        form = BatchForm(request.POST)
        
        if form.is_valid():
            try:
                batch = form.save(commit=True)
                messages.success(request, f'Batch "{batch.batch_code}" created successfully')
                
                # Return JavaScript to close window
                return HttpResponse("""
                    <script>
                        if (window.opener && !window.opener.closed) {
                            window.opener.location.reload();
                        }
                        window.close();
                    </script>
                """)
            except Exception as e:
                messages.error(request, f'Error creating batch: {str(e)}')
        else:
            # Collect all form errors
            error_messages = []
            for field, errors in form.errors.items():
                field_name = field.replace('_', ' ').title()
                for error in errors:
                    error_messages.append(f"{field_name}: {error}")
            
            if error_messages:
                messages.error(request, '<br>'.join(error_messages))
    
    else:
        form = BatchForm()
    
    return render(request, 'batch_create.html', {'form': form})


@login_required
@has_a_auhtenticated_user
def update_batch(request, pk):
    """Update existing batch (opens in modal window)"""
    batch = get_object_or_404(Batch, pk=pk)
    
    if request.method == 'POST':
        form = BatchForm(request.POST, instance=batch)
        
        if form.is_valid():
            try:
                batch = form.save(commit=True)
                messages.success(request, f'Batch "{batch.batch_code}" updated successfully')
                
                # Return JavaScript to close window
                return HttpResponse("""
                    <script>
                        if (window.opener && !window.opener.closed) {
                            window.opener.location.reload();
                        }
                        window.close();
                    </script>
                """)
            except Exception as e:
                messages.error(request, f'Error updating batch: {str(e)}')
        else:
            # Collect all form errors
            error_messages = []
            for field, errors in form.errors.items():
                field_name = field.replace('_', ' ').title()
                for error in errors:
                    error_messages.append(f"{field_name}: {error}")
            
            if error_messages:
                messages.error(request, '<br>'.join(error_messages))
    
    else:
        form = BatchForm(instance=batch)
    
    return render(request, 'batch_update.html', {'form': form, 'batch': batch})


@login_required
@has_a_auhtenticated_user
def batch_detail(request, pk):
    """View batch details with enrolled students"""
    batch = get_object_or_404(Batch, pk=pk)
    
    # Get enrolled students
    enrollments = batch.enrollments.select_related('student').order_by('-enrollment_date')
    
    # Batch statistics
    total_students = batch.total_students
    current_students = batch.current_students
    completed_students = batch.enrollments.filter(status='completed').count()
    active_students = batch.enrollments.filter(status='active').count()
    
    # Payment statistics
    payments = []
    total_paid = Decimal('0.00')
    total_due = Decimal('0.00')
    
    for enrollment in enrollments:
        paid_amount = enrollment.paid_amount
        due_amount = enrollment.balance
        payments.append({
            'enrollment': enrollment,
            'paid': paid_amount,
            'due': due_amount
        })
        total_paid += paid_amount
        total_due += due_amount
    
    # Get course details
    course = batch.course
    course_details = {}
    if course:
        course_details = {
            'title': getattr(course, 'title', 'N/A'),
            'duration': getattr(course, 'duration', 'N/A'),
            'original_price': getattr(course, 'original_price', 'N/A'),
            'offer_price': getattr(course, 'offer_price', 'N/A'),
            'rating': getattr(course, 'rating', 'N/A'),
        }
    
    context = {
        'batch': batch,
        'enrollments': enrollments,
        'payments': payments,
        'total_students': total_students,
        'current_students': current_students,
        'completed_students': completed_students,
        'active_students': active_students,
        'total_paid': total_paid,
        'total_due': total_due,
        'course_details': course_details,
        'available_seats': batch.available_seats,
        'occupancy_percentage': batch.occupancy_percentage,
        'duration_days': batch.duration_days,
        'days_remaining': batch.days_remaining,
    }
    
    return render(request, 'batch_details.html', context)


@login_required
@has_a_auhtenticated_user
def delete_batch(request, pk):
    """Delete batch with confirmation"""
    batch = get_object_or_404(Batch, pk=pk)
    
    # Check if batch has enrollments
    if batch.enrollments.exists():
        messages.error(request, f'Cannot delete batch "{batch.batch_code}" because it has enrolled students.')
        return redirect(request.META.get('HTTP_REFERER', '/software/batches/'))
    
    if batch:
        batch_code = batch.batch_code
        batch.delete()
        messages.success(request, f'Batch "{batch_code}" deleted successfully')
        return redirect(request.META.get('HTTP_REFERER', '/software/batches/'))
    
    messages.info(request, 'Nothing to delete.')
    # For GET request, should be handled by JavaScript modal
    return redirect(request.META.get('HTTP_REFERER', '/software/batches/'))

 
# For better Excel formatting, you can also create an XLSX version:
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


@login_required
@has_a_auhtenticated_user
def export_batch_report_excel(request, batch_id):
    """Export batch report as Excel (XLSX) with advanced formatting"""
    from django.http import HttpResponse
    import tempfile
    import os
    
    batch = get_object_or_404(Batch, id=batch_id)
    
    # Calculate statistics (same as CSV version)
    total_students = batch.enrollments.count()
    active_students = batch.enrollments.filter(status='active').count()
    payment_completed_students = batch.enrollments.filter(
        payment_status='completed'
    ).count()
    
    total_collected = batch.enrollments.aggregate(
        total_collected=Sum('payments__amount')
    )['total_collected'] or Decimal('0.00')
    
    total_pending = Decimal('0.00')
    for enrollment in batch.enrollments.all():
        total_pending += enrollment.balance
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Batch Report"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    title_font = Font(bold=True, size=14, color="1F4E78")
    subtitle_font = Font(bold=True, size=12, color="2F75B5")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 1. BATCH INFORMATION SECTION
    ws.merge_cells('A1:E1')
    ws['A1'] = "BATCH REPORT"
    ws['A1'].font = Font(bold=True, size=16, color="1F4E78")
    ws['A1'].alignment = Alignment(horizontal="center")
    
    ws['A3'] = "Batch Information"
    ws['A3'].font = title_font
    
    info_rows = [
        ["Batch Code:", batch.batch_code],
        ["Batch Title:", batch.batch_title],
        ["Course Type:", batch.course_type],
        ["Course:", batch.course.title if batch.course else 'N/A'],
        ["Start Date:", batch.start_date],
        ["End Date:", batch.end_date if batch.end_date else 'Ongoing'],
        ["Batch Status:", batch.batch_status.title()],
        ["Max Students:", batch.max_students],
        ["Duration:", f"{batch.duration_days} days" if batch.duration_days else "N/A"],
    ]
    
    for i, (label, value) in enumerate(info_rows, start=4):
        ws[f'A{i}'] = label
        ws[f'A{i}'].font = Font(bold=True)
        ws[f'B{i}'] = value
    
    # 2. STATISTICS SECTION
    stat_row = 14
    ws[f'A{stat_row}'] = "Batch Statistics"
    ws[f'A{stat_row}'].font = title_font
    
    # Statistics headers
    stat_headers = ["Metric", "Count", "Percentage"]
    for col, header in enumerate(stat_headers, start=1):
        cell = ws.cell(row=stat_row+1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Statistics data
    stats_data = [
        ["Total Enrolled Students", total_students, "100%"],
        ["Active Students", active_students, 
         f"{(active_students/total_students*100):.1f}%" if total_students > 0 else "0%"],
        ["Payment Completed Students", payment_completed_students,
         f"{(payment_completed_students/total_students*100):.1f}%" if total_students > 0 else "0%"],
        ["Available Seats", max(batch.max_students - total_students, 0),
         f"Occupancy: {(total_students/batch.max_students*100):.1f}%" if batch.max_students > 0 else "N/A"],
    ]
    
    for i, row_data in enumerate(stats_data, start=stat_row+2):
        for j, value in enumerate(row_data, start=1):
            cell = ws.cell(row=i, column=j)
            cell.value = value
            cell.border = border
    
    # 3. FINANCIAL SUMMARY SECTION
    financial_row = stat_row + len(stats_data) + 3
    ws[f'A{financial_row}'] = "Financial Summary"
    ws[f'A{financial_row}'].font = title_font
    
    # Financial headers
    fin_headers = ["Description", "Amount (₹)"]
    for col, header in enumerate(fin_headers, start=1):
        cell = ws.cell(row=financial_row+1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.alignment = header_alignment
        cell.border = border
    
    # Financial data
    total_potential_fees = batch.enrollments.aggregate(total=Sum('total_fees'))['total'] or Decimal('0.00')
    total_discounts = batch.enrollments.aggregate(total=Sum('discount'))['total'] or Decimal('0.00')
    net_fees_total = batch.enrollments.aggregate(total=Sum('net_fees'))['total'] or Decimal('0.00')
    collection_percentage = (total_collected / net_fees_total * 100) if net_fees_total > 0 else 0
    
    fin_data = [
        ["Total Potential Fees", f"₹{total_potential_fees:,.2f}"],
        ["Total Discounts Given", f"₹{total_discounts:,.2f}"],
        ["Total Net Fees", f"₹{net_fees_total:,.2f}"],
        ["Total Collected Amount", f"₹{total_collected:,.2f}"],
        ["Total Pending Amount", f"₹{total_pending:,.2f}"],
        ["Collection Rate", f"{collection_percentage:.1f}%"],
    ]
    
    for i, row_data in enumerate(fin_data, start=financial_row+2):
        for j, value in enumerate(row_data, start=1):
            cell = ws.cell(row=i, column=j)
            cell.value = value
            cell.border = border
    
    # 4. STUDENT DETAILS SECTION
    student_row = financial_row + len(fin_data) + 3
    ws[f'A{student_row}'] = "STUDENT DETAILS"
    ws[f'A{student_row}'].font = title_font
    
    # Student headers
    student_headers = [
        "S.No.", "Enrollment ID", "Student Name", "Mobile Number", 
        "Enrollment Date", "Status", "Payment Status",
        "Total Fees (₹)", "Discount (₹)", "Net Fees (₹)", 
        "Paid Amount (₹)", "Balance (₹)", "Payment Progress"
    ]
    
    for col, header in enumerate(student_headers, start=1):
        cell = ws.cell(row=student_row+1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        cell.alignment = header_alignment
        cell.border = border
    
    # Student data
    enrollments = batch.enrollments.select_related('student').all()
    for i, enrollment in enumerate(enrollments, start=student_row+2):
        row_data = [
            i - student_row - 1,
            enrollment.enrollment_id,
            enrollment.student.name,
            enrollment.student.mobile_number,
            enrollment.enrollment_date.strftime('%d-%b-%Y'),
            enrollment.status.title(),
            enrollment.payment_status.title(),
            enrollment.total_fees,
            enrollment.discount,
            enrollment.net_fees,
            enrollment.paid_amount,
            enrollment.balance,
            enrollment.payment_progress
        ]
        
        for j, value in enumerate(row_data, start=1):
            cell = ws.cell(row=i, column=j)
            cell.value = value
            cell.border = border
    
    # 5. PAYMENT DETAILS SECTION
    payment_row = student_row + len(enrollments) + 3
    ws[f'A{payment_row}'] = "PAYMENT TRANSACTIONS"
    ws[f'A{payment_row}'].font = title_font
    
    # Payment headers
    payment_headers = [
        "Student Name", "Payment Date", "Payment ID", "Type", 
        "Mode", "Amount (₹)", "Reference", "Received By"
    ]
    
    for col, header in enumerate(payment_headers, start=1):
        cell = ws.cell(row=payment_row+1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
        cell.alignment = header_alignment
        cell.border = border
    
    # Payment data
    payments = Payment.objects.filter(
        enrollment__batch=batch
    ).select_related('enrollment__student', 'received_by').order_by('-payment_date')
    
    for i, payment in enumerate(payments, start=payment_row+2):
        row_data = [
            payment.enrollment.student.name,
            payment.payment_date.strftime('%d-%b-%Y'),
            payment.payment_id,
            payment.get_payment_type_display(),
            payment.get_payment_mode_display(),
            payment.amount,
            payment.reference_number or '-',
            payment.received_by.name if payment.received_by else '-'
        ]
        
        for j, value in enumerate(row_data, start=1):
            cell = ws.cell(row=i, column=j)
            cell.value = value
            cell.border = border
    
    # Adjust column widths
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2
    
    # Save to temporary file
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        wb.save(tmp.name)
        tmp_path = tmp.name
    
    # Create response
    with open(tmp_path, 'rb') as f:
        response = HttpResponse(
            f.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="batch_report_{batch.batch_code}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    # Clean up temp file
    os.unlink(tmp_path)
    
    return response

  


# views.py
from django.http import HttpResponse
from django.template.loader import get_template
from django.conf import settings
from decimal import Decimal
import os
from datetime import datetime


@login_required
@has_a_auhtenticated_user
def print_payment_receipt(request, pk):
    """Generate printable receipt for a payment"""
    payment = get_object_or_404(Payment, pk=pk)
    enrollment = payment.enrollment
    
    # Get institute information (you can store these in settings or database)
    institute_info = {
        'name': 'VR Training Academy',
        'address': 'Airport Rd, near Kimaya Garden, Shirdi, Maharashtra 423109',
        'email': 'vrtrainingacademy1123@gmail.com',
        'primary_contact': '+91 99703 60424',
        'whatsapp_contact': '+91 92702 01760',
    }
    
    # Calculate receipt details
    receipt_data = {
        'payment': payment,
        'enrollment': enrollment,
        'student': enrollment.student,
        'batch': enrollment.batch,
        'institute': institute_info,
        'receipt_date': timezone.now().date(),
        'receipt_number': f"RCPT-{payment.payment_id}",
        
        # Course details
        'course_title': enrollment.course_title or "Not specified",
        'batch_code': enrollment.batch.batch_code if enrollment.batch else "Not assigned",
        
        # Payment summary
        'total_fees': enrollment.total_fees,
        'discount': enrollment.discount,
        'net_fees': enrollment.net_fees,
        'paid_before': enrollment.paid_amount - payment.amount,  # Amount paid before this payment
        'paid_after': enrollment.paid_amount,  # Total paid including this payment
        'balance_after': enrollment.balance,  # Balance after this payment
        'paid_in_words': amount_in_words(payment.amount),
    }
    
    # For PDF download
    download = request.GET.get('download', False)
    
    if download:
        # Generate PDF (requires reportlab or similar library)
        # For now, we'll just return HTML that users can print
        return generate_pdf_receipt(receipt_data)
    
    context = {
        **receipt_data,
        'download_mode': download,
    }
    
    return render(request, 'print_payment_receipt.html', context)

def amount_in_words(amount):
    """Convert amount to words (basic implementation)"""
    try:
        # Simple implementation - you might want a more robust one
        num = int(amount)
        
        units = ["", "One", "Two", "Three", "Four", "Five", "Six", "Seven", "Eight", "Nine"]
        teens = ["Ten", "Eleven", "Twelve", "Thirteen", "Fourteen", "Fifteen", "Sixteen", 
                "Seventeen", "Eighteen", "Nineteen"]
        tens = ["", "", "Twenty", "Thirty", "Forty", "Fifty", "Sixty", "Seventy", "Eighty", "Ninety"]
        
        if num == 0:
            return "Zero"
        
        words = ""
        
        # Handle thousands
        if num >= 1000:
            thousands = num // 1000
            if thousands > 0:
                words += units[thousands] + " Thousand "
            num %= 1000
        
        # Handle hundreds
        if num >= 100:
            hundreds = num // 100
            words += units[hundreds] + " Hundred "
            num %= 100
        
        # Handle tens and units
        if num >= 20:
            tens_digit = num // 10
            words += tens[tens_digit] + " "
            num %= 10
        
        if num >= 10:
            words += teens[num - 10] + " "
            num = 0
        
        if num > 0:
            words += units[num] + " "
        
        words += "Rupees Only"
        return words.strip()
    
    except:
        return f"{amount} Rupees Only"


@login_required
@has_a_auhtenticated_user
def generate_pdf_receipt(receipt_data):
    """Generate PDF receipt (optional - implement if you have PDF library)"""
    # This is a placeholder - you can implement PDF generation using:
    # 1. ReportLab
    # 2. WeasyPrint
    # 3. xhtml2pdf
    # For now, we'll return HTML that users can print
    
    template = get_template('print_payment_receipt.html')
    html = template.render({'receipt_data': receipt_data, 'download_mode': True})
    
    # Create PDF response
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Receipt_{receipt_data["receipt_number"]}.pdf"'
    
    # Here you would convert HTML to PDF using your preferred library
    # For example with xhtml2pdf:
    # from xhtml2pdf import pisa
    # pisa.CreatePDF(html, dest=response)
    
    return response










@login_required
def event_list(request):
    """
    List all events with filtering, pagination, and statistics
    Follows the exact pattern of crm_follow_up_list
    """
    # Get filter parameters
    search_query = request.GET.get('search', '')
    status_filter = request.GET.get('status', '')
    category_filter = request.GET.get('category', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    # Start with all events
    events = Event.objects.all().order_by('-event_date', '-created_at')
    
    # Apply filters
    if search_query:
        events = events.filter(
            Q(title__icontains=search_query) |
            Q(subtitle__icontains=search_query) |
            Q(description__icontains=search_query) |
            Q(category__icontains=search_query)
        )
    
    if status_filter:
        events = events.filter(status=status_filter)
    
    if category_filter:
        events = events.filter(category__icontains=category_filter)
    
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d')
            events = events.filter(event_date__gte=date_from_obj)
        except ValueError:
            pass
    
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d')
            events = events.filter(event_date__lte=date_to_obj)
        except ValueError:
            pass
    
    # Calculate statistics
    total_events = events.count()
    
    # Upcoming events (future dates)
    today = date.today()
    upcoming_events = events.filter(event_date__gte=today).count()
    
    # Total registrations
    total_registrations = EventRegistration.objects.filter(
        event__in=events
    ).count()
    
    # Total revenue from paid events
    total_revenue = EventRegistration.objects.filter(
        event__in=events,
        payment_status='success'
    ).aggregate(total=Sum('amount_paid'))['total'] or Decimal('0.00')
    
    # Add calculated properties to each event for template
    for event in events:
        # Registrations count
        event.registrations_count = event.registrations.count()
        
        # Successful registrations (paid)
        event.successful_registrations = event.registrations.filter(
            payment_status='success'
        ).count()
        
        # Pending registrations
        event.pending_registrations = event.registrations.filter(
            payment_status='pending'
        ).count()
        
        # Total revenue for this event
        event.event_revenue = event.registrations.filter(
            payment_status='success'
        ).aggregate(total=Sum('amount_paid'))['total'] or Decimal('0.00')
        
        # Pending revenue (if any pending payments)
        event.pending_revenue = event.registrations.filter(
            payment_status='pending'
        ).aggregate(total=Sum('amount_paid'))['total'] or Decimal('0.00')
        
        # Total potential revenue (if all registrations were paid)
        if event.is_free:
            event.total_potential_revenue = Decimal('0.00')
        else:
            # Calculate potential revenue: registration count * event price
            potential_count = event.registrations_count
            event.total_potential_revenue = potential_count * event.registration_offer_fee
        
        # Check if event is upcoming
        event.is_upcoming = event.event_date >= today
        
        # Check if event is today
        event.is_today = event.event_date == today
    
    # Get unique categories for filter dropdown
    categories = Event.objects.exclude(category__isnull=True).exclude(category='').values_list(
        'category', flat=True
    ).distinct().order_by('category')
    
    # Pagination (10 items per page, same as follow-ups)
    paginator = Paginator(events, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Prepare context (similar to crm_follow_up_list)
    context = {
        'events': page_obj,
        'categories': categories,
        
        # Statistics
        'total_events': total_events,
        'upcoming_events': upcoming_events,
        'total_registrations': total_registrations,
        'total_revenue': total_revenue,
        
        # Filter values for form preservation
        'search_query': search_query,
        'status_filter': status_filter,
        'category_filter': category_filter,
        'date_from': date_from,
        'date_to': date_to,
        
        # Status choices for filter dropdown
        'status_choices': Event.EVENT_STATUS,
    }
    
    return render(request, 'enrollment_event_list.html', context)



@login_required
def event_detail(request, pk):
    """View event details with registrations"""
    event = get_object_or_404(Event, pk=pk)
    
    # Get registrations for this event
    registrations = event.registrations.all().order_by('-created_at')
    
    # Calculate registration stats
    total_registrations = registrations.count()
    successful_registrations = registrations.filter(payment_status='success').count()
    pending_registrations = registrations.filter(payment_status='pending').count()
    
    # Calculate revenue
    total_revenue = registrations.filter(payment_status='success').aggregate(
        total=Sum('amount_paid')
    )['total'] or Decimal('0.00')
    
    # Paginate registrations
    paginator = Paginator(registrations, 10)
    page_number = request.GET.get('page')
    registrations_page = paginator.get_page(page_number)
    
    context = {
        'event': event,
        'registrations': registrations_page,
        'total_registrations': total_registrations,
        'successful_registrations': successful_registrations,
        'pending_registrations': pending_registrations,
        'total_revenue': total_revenue,
    }
    
    return render(request, 'enrollment_event_detail.html', context)

@login_required
def create_event(request):
    """Create new event (opens in modal window)"""
    if request.method == 'POST':
        form = EventForm(request.POST, request.FILES)
        
        if form.is_valid():
            try:
                event = form.save(commit=True)
                messages.success(request, f'Event "{event.title}" created successfully')
                
                # Return JavaScript to close window
                return HttpResponse("""
                    <script>
                        if (window.opener && !window.opener.closed) {
                            window.opener.location.reload();
                        }
                        window.close();
                    </script>
                """)
            except Exception as e:
                messages.error(request, f'Error creating event: {str(e)}')
        else:
            # Collect all form errors
            error_messages = []
            for field, errors in form.errors.items():
                field_name = field.replace('_', ' ').title()
                for error in errors:
                    error_messages.append(f"{field_name}: {error}")
            
            if error_messages:
                messages.error(request, '<br>'.join(error_messages))
    
    else:
        form = EventForm()
    
    return render(request, 'enrollment_event_create.html', {'form': form})

@login_required
def update_event(request, pk):
    """Update existing event (opens in modal window)"""
    event = get_object_or_404(Event, pk=pk)
    
    if request.method == 'POST':
        form = EventForm(request.POST, request.FILES, instance=event)
        
        if form.is_valid():
            try:
                event = form.save(commit=True)
                messages.success(request, f'Event "{event.title}" updated successfully')
                
                # Return JavaScript to close window
                return HttpResponse("""
                    <script>
                        if (window.opener && !window.opener.closed) {
                            window.opener.location.reload();
                        }
                        window.close();
                    </script>
                """)
            except Exception as e:
                messages.error(request, f'Error updating event: {str(e)}')
        else:
            # Collect all form errors
            error_messages = []
            for field, errors in form.errors.items():
                field_name = field.replace('_', ' ').title()
                for error in errors:
                    error_messages.append(f"{field_name}: {error}")
            
            if error_messages:
                messages.error(request, '<br>'.join(error_messages))
    
    else:
        form = EventForm(instance=event)
    
    return render(request, 'enrollment_event_update.html', {'form': form, 'event': event})

@login_required
def delete_event(request, pk):
    """Delete event with confirmation"""
    event = get_object_or_404(Event, pk=pk)
    
    event_title = event.title
    event.delete()
    messages.success(request, f'Event "{event_title}" deleted successfully')
    return redirect(request.META.get('HTTP_REFERER', 'enrollment-events'))    
     
 
 
from decimal import Decimal
import num2words
# views.py
from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse, Http404
from django.template.loader import render_to_string
import num2words
import logging

logger = logging.getLogger(__name__)

def print_event_receipt(request, registration_id):
    """
    View to print event registration receipt
    """
    try:
        # Try to find by registration_id field first
        registration = EventRegistration.objects.get(registration_id=registration_id)
    except EventRegistration.DoesNotExist:
        try:
            # If not found, try by primary key (id)
            registration = EventRegistration.objects.get(id=registration_id)
        except (EventRegistration.DoesNotExist, ValueError):
            # Log the error for debugging
            logger.error(f"Registration not found: {registration_id}")
            raise Http404(f"No registration found with ID: {registration_id}")
    
    event = registration.event
    
    # Convert amount to words
    try:
        if event.is_free:
            amount_in_words = "Free"
        else:
            amount_paid = float(registration.amount_paid or 0)
            amount_in_words = num2words.num2words(
                amount_paid,
                lang='en_IN'
            ).title() + " Rupees"
    except Exception as e:
        logger.error(f"Error converting amount to words: {e}")
        amount_in_words = "Zero Rupees"
    
    context = {
        'registration': registration,
        'event': event,
        'amount_in_words': amount_in_words,
    }
    
    # Check if autoprint is requested
    if request.GET.get('autoprint', 'false') == 'true':
        context['autoprint'] = True
    
    return render(request, 'print_event_registration_receipt.html', context)

def download_event_receipt(request, registration_id):
    """
    Download HTML receipt file
    """
    try:
        registration = EventRegistration.objects.get(registration_id=registration_id)
    except EventRegistration.DoesNotExist:
        registration = get_object_or_404(EventRegistration, id=registration_id)
    
    event = registration.event
    
    # Convert amount to words
    if event.is_free:
        amount_in_words = "Free"
    else:
        amount_paid = float(registration.amount_paid or 0)
        amount_in_words = num2words.num2words(
            amount_paid,
            lang='en_IN'
        ).title() + " Rupees"
    
    context = {
        'registration': registration,
        'event': event,
        'amount_in_words': amount_in_words,
    }
    
    # Render HTML
    html_string = render_to_string('print_event_receipt.html', context)
    
    # Return as downloadable HTML file
    response = HttpResponse(html_string, content_type='text/html')
    response['Content-Disposition'] = f'attachment; filename="Receipt_{registration.registration_id}.html"'
    return response

def event_registration_list(request):
    """
    List all registrations for debugging/management
    """
    registrations = EventRegistration.objects.select_related('event').all()
    
    context = {
        'registrations': registrations
    }
    return render(request, 'registration_list.html', context)